#!/usr/bin/env python3
"""
AMF1 Partner Survey - Phase 0 Reference Document Builder

Parses global table (Banner) Excel files for Wave 1 and Wave 2.
Extracts expected values for all in-scope dashboard questions at:
  - Total sample level
  - Each individual filter slice (country, age, gender, F1 fan, etc.)

Output: outputs/reference_document.csv  (ground truth for Phase 1 verification)

Usage: python scripts/build_reference_doc.py
"""

import csv
import re
import sys
import io
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent.parent

FILES = {
    "W1": {
        "Banner1": BASE_DIR / "data" / "Wave1" / "Banner1.xlsx",
        "Banner2": BASE_DIR / "data" / "Wave1" / "Banner2.xlsx",
        "Banner3": BASE_DIR / "data" / "Wave1" / "Banner3.xlsx",
    },
    "W2": {
        "Banner1": BASE_DIR / "data" / "Wave2" / "Banner1.xlsx",
        "Banner2": BASE_DIR / "data" / "Wave2" / "Banner2.xlsx",
        "Banner3": BASE_DIR / "data" / "Wave2" / "Banner3.xlsx",
    },
}

OUTPUT_CSV = BASE_DIR / "outputs" / "reference_document.csv"

# ── In-scope / out-of-scope question filters ───────────────────────────────────
# Match on the leading question code prefix (case-insensitive)
IN_SCOPE_PREFIXES = {
    # Overview P0
    "D1", "D2", "D3", "D4", "D5", "D6",
    # Routing / partner questions used in dashboard
    "D7", "D8", "D9",
    "D14", "D15", "D16",
    # Partner-specific
    "B1", "B2", "B2A", "B3", "B5", "B7",
    "C5", "C6",
    "ARC1", "ARC2", "ARC5", "ARC6", "ARC7", "ARC8",
    "CON1", "CON2", "CON3", "CON4", "CON5", "CON6", "CON7", "CON8",
    "COR1",
    "MAA1", "MAA2", "MAA3", "MAA4", "MAA5", "MAA6", "MAA7", "MAA8",
    "VAL1",
    "PEP1", "PEP2",
    "NEX1",
    "GLN1",
    "SER1",
    "ATL1", "ATL2",
    "COG1", "COG2", "COG3",
    "TTK1", "TTK2", "TTK3", "TTK4", "TTK5", "TTK6", "TTK7",
}

# Explicitly excluded (out of scope per brief)
OUT_OF_SCOPE_PREFIXES = {"D10", "D11", "D12", "D13"}

# Skip these generic response row labels (non-data rows)
SKIP_LABELS = {
    "sigma", "mean", "std. dev.", "std. err.", "median", "base : all respondents",
    "base :", "proportions/means:", "* small base", "** very small base",
}


def extract_question_code(text: str) -> str:
    """Pull leading question code from question text, e.g. 'D1. ...' → 'D1'."""
    if not text:
        return ""
    m = re.match(r"^([A-Za-z]+\d+[a-zA-Z]?)\b", text.strip())
    return m.group(1).upper() if m else ""


def is_in_scope(qcode: str) -> bool:
    if not qcode:
        return False
    uc = qcode.upper()
    if uc in OUT_OF_SCOPE_PREFIXES:
        return False
    # Exact match first
    if uc in IN_SCOPE_PREFIXES:
        return True
    # Prefix match: e.g. "D14" matches "D14" prefix
    for pfx in IN_SCOPE_PREFIXES:
        if uc.startswith(pfx):
            return True
    return False


def clean_numeric(val):
    """Convert cell value to float, handling *, **, '- ' strings."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace("*", "").replace("- ", "").replace("-", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def clean_base_n(val):
    """Convert base n cell to int, returning None for very small bases."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val)
    stars = s.count("*")
    num_s = s.replace("*", "").strip()
    try:
        n = int(float(num_s))
        return n  # keep even if small; Phase 1 handles low-n warnings
    except ValueError:
        return None


def parse_summary_sheet(ws) -> dict:
    """Returns {table_num (int): question_text (str)}."""
    mapping = {}
    for row in ws.iter_rows(values_only=True):
        if len(row) < 2 or not row[0] or not row[1]:
            continue
        m = re.match(r"^Table\s+(\d+)$", str(row[0]).strip())
        if m:
            mapping[int(m.group(1))] = str(row[1]).strip()
    return mapping


def parse_t1_sheet(ws, summary_map: dict):
    """
    Generator: yields one dict per (table × response_row × filter_col).
    """
    # Load all rows into a list for index-based access
    all_rows = list(ws.iter_rows(values_only=True))

    # Locate table start positions
    table_positions = []
    for i, row in enumerate(all_rows):
        if row[0] and re.match(r"^Table\s+\d+$", str(row[0]).strip()):
            m = re.match(r"^Table\s+(\d+)$", str(row[0]).strip())
            if m:
                table_positions.append((i, int(m.group(1))))

    n_tables = len(table_positions)

    for t_idx, (start_i, table_num) in enumerate(table_positions):
        end_i = table_positions[t_idx + 1][0] if t_idx + 1 < n_tables else len(all_rows)
        trows = all_rows[start_i:end_i]

        if len(trows) < 10:
            continue

        # Resolve question text via summary map (more reliable than sheet text)
        question_text = summary_map.get(table_num, "")
        if not question_text and trows[2][0]:
            question_text = str(trows[2][0]).strip()

        qcode = extract_question_code(question_text)
        if not is_in_scope(qcode):
            continue

        # Dynamically locate header block (handles multi-line question text)
        # Find the group row by looking for: row[1] == 'Total' (or similar group marker)
        # This is robust to 1-line or 2-line question text.
        group_row_idx = None
        for hi in range(2, min(8, len(trows))):
            r = trows[hi]
            if (r[0] is None and len(r) > 1 and r[1] is not None
                    and str(r[1]).strip() in ("Total",)):
                group_row_idx = hi
                break

        if group_row_idx is None:
            # Fallback to fixed offset
            group_row_idx = 3

        # Based on group row, locate subsequent header rows
        # group_row (hi), blank (hi+1), filter_names (hi+2), letters (hi+3), base (hi+4)
        # but sometimes blank is missing: filter_names (hi+1), letters (hi+2), base (hi+3)
        # Detect by checking if hi+1 is blank
        def _is_blank(row):
            return all(v is None or str(v).strip() == "" for v in row)

        hi = group_row_idx
        if hi + 4 < len(trows):
            if _is_blank(trows[hi + 1]):
                filter_row_idx = hi + 2
                letter_row_idx = hi + 3
                base_row_idx   = hi + 4
            else:
                filter_row_idx = hi + 1
                letter_row_idx = hi + 2
                base_row_idx   = hi + 3
        else:
            filter_row_idx = hi + 2
            letter_row_idx = hi + 3
            base_row_idx   = hi + 4

        if base_row_idx >= len(trows):
            continue

        group_row  = trows[group_row_idx]
        filter_row = trows[filter_row_idx] if filter_row_idx < len(trows) else ()
        letter_row = trows[letter_row_idx] if letter_row_idx < len(trows) else ()
        base_row   = trows[base_row_idx]
        data_start = base_row_idx + 2  # skip base + blank spacer

        # Build column map: list of (group_label, filter_label, col_letter, base_n)
        # Index 0 = label column (skip); index 1 = Total; index 2+ = filter slices
        num_cols = len(filter_row)
        col_meta = []  # one entry per data column (starting at index 1)
        current_group = "Total"
        for c in range(1, num_cols):
            # Update current group when a new group header appears
            if (c < len(group_row) and group_row[c] is not None
                    and str(group_row[c]).strip()):
                current_group = str(group_row[c]).strip()

            if c == 1:
                grp = "Total"
                flt = "Total"
                ltr = "Total"
            else:
                grp = current_group
                flt = str(filter_row[c]).strip() if c < len(filter_row) and filter_row[c] else ""
                ltr = str(letter_row[c]).strip() if c < len(letter_row) and letter_row[c] else ""

            bn = clean_base_n(base_row[c] if c < len(base_row) else None)
            col_meta.append((grp, flt, ltr, bn))

        # Parse data rows (triplets: count row, pct row, significance row)
        i = data_start
        while i < len(trows):
            row = trows[i]

            # End of data section
            if row[0] and str(row[0]).startswith("____"):
                break

            label = str(row[0]).strip() if row[0] is not None else ""
            label_lower = label.lower()

            # Skip blank / utility rows
            if not label or label in (" ", ""):
                i += 1
                continue
            if any(label_lower.startswith(s) for s in SKIP_LABELS):
                i += 1
                continue
            # Skip statistical summaries that aren't question responses
            if label in ("Sigma", "Mean", "Std. Dev.", "Std. Err.", "Median"):
                i += 1
                continue

            # The next row should be the percentage/value row (col A = None)
            pct_row = trows[i + 1] if i + 1 < len(trows) else None

            # Detect NPS-style tables where values are in the label row itself
            # (no separate pct row — next row is sig codes or another base row)
            values_in_label_row = (
                pct_row is None
                or pct_row[0] is not None
                or all(
                    pct_row[c] is None or str(pct_row[c]).strip() in ("", " ")
                    or (isinstance(pct_row[c], str) and not pct_row[c].replace(".", "").replace("-", "").replace("*", "").strip().isdigit())
                    for c in range(1, min(5, len(pct_row)))
                )
            )

            if values_in_label_row and pct_row is not None and pct_row[0] is not None:
                # pct_row col A is non-None → not a pct row at all, skip this label
                i += 1
                continue

            # Emit one record per filter column
            for c_idx, (grp, flt, ltr, bn) in enumerate(col_meta):
                excel_col = c_idx + 1  # offset from col A (label col)

                raw_count = row[excel_col] if excel_col < len(row) else None
                raw_pct   = (pct_row[excel_col] if pct_row and excel_col < len(pct_row) else None)

                count = clean_numeric(raw_count)
                pct   = clean_numeric(raw_pct)

                # For NPS/score tables: value is in the count row (decimal or integer score)
                if values_in_label_row and pct is None and count is not None:
                    pct = count
                    count = None

                # Convert proportion to percentage where applicable
                if pct is not None and -1.0 <= pct <= 1.0:
                    pct_display = round(pct * 100, 1)
                    value_type  = "percentage"
                else:
                    pct_display = pct  # NPS score, mean, etc. — keep raw
                    value_type  = "score_or_mean"

                yield {
                    "wave":           "",
                    "banner":         "",
                    "table_num":      table_num,
                    "question_code":  qcode,
                    "question_text":  question_text,
                    "response_label": label,
                    "filter_group":   grp,
                    "filter_option":  flt,
                    "col_letter":     ltr,
                    "base_n":         bn,
                    "count":          int(count) if count is not None and value_type == "percentage" else count,
                    "value":          pct_display,
                    "value_type":     value_type,
                }

            # Advance past count + pct + sig rows (or count + sig for NPS-style)
            i += 3


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    records = []
    missing_files = []

    for wave, banners in FILES.items():
        for banner, filepath in banners.items():
            if not filepath.exists():
                missing_files.append(str(filepath))
                print(f"  MISSING: {filepath}", file=sys.stderr)
                continue

            print(f"Processing {wave} {banner}: {filepath.name} ...", file=sys.stderr)
            try:
                wb = openpyxl.load_workbook(str(filepath), read_only=True)
            except Exception as e:
                print(f"  ERROR loading: {e}", file=sys.stderr)
                continue

            summary_map = parse_summary_sheet(wb["Summary"])
            batch = 0
            for rec in parse_t1_sheet(wb["T1"], summary_map):
                rec["wave"]   = wave
                rec["banner"] = banner
                records.append(rec)
                batch += 1

            wb.close()
            print(f"  → {batch} records extracted", file=sys.stderr)

    print(f"\nTotal records: {len(records)}", file=sys.stderr)

    if not records:
        print("No records extracted. Check file paths and in-scope filters.", file=sys.stderr)
        return

    fieldnames = [
        "wave", "banner", "table_num", "question_code", "question_text",
        "response_label", "filter_group", "filter_option", "col_letter",
        "base_n", "count", "value", "value_type",
    ]
    with open(str(OUTPUT_CSV), "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)

    print(f"\nReference document: {OUTPUT_CSV}", file=sys.stderr)
    print("Run 'python scripts/build_ambiguity_report.py' to regenerate ambiguities report.", file=sys.stderr)


if __name__ == "__main__":
    main()
